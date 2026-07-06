import fitz  # PyMuPDF
from docx import Document 
import csv
from openpyxl import load_workbook
import easyocr
import textract

def read_pdf(file_path):
    text = ""

    doc = fitz.open(file_path)

    for page in doc:
        text += page.get_text()

    doc.close()

    return text

def read_docx(file_path):

    text = ""

    doc = Document(file_path)

    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"

    return text



def read_excel(file_path):
    workbook = load_workbook(file_path)

    text = ""

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    text += str(cell) + " "

            text += "\n"

    return text


def read_csv(file_path):
    text = ""

    with open(file_path, "r", encoding="utf-8") as file:
        reader = csv.reader(file)

        for row in reader:
            text += " ".join(row) + "\n"

    return text

def read_txt(file_path):
    with open(file_path, "r", encoding="utf-8") as file:
        return file.read()
    
    
def read_doc(file_path):

    text = textract.process(file_path)

    return text.decode("utf-8")



